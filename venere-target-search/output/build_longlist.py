# -*- coding: utf-8 -*-
"""Genera la longlist consolidata Progetto Venere (medicina estetica IT) in xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colonne: Tier, Cluster, Regione, Clinica, Citta, Prov, Sito, Sedi, Medici,
# Titolare, Servizi, Angolo, Segnali di fit, Fonte
# Angolo: MP=multi-provider gia strutturato | GEN=uscita generazionale |
#         PLAT=piattaforma/multi-sede bolt-on | CONV=mono/convertibile
T = "targets"
rows = [
 # ---------------- LOMBARDIA (Nord-Ovest) ----------------
 ["A","Lombardia (NO)","Lombardia","Poliambulatorio Medico Chirurgico Filippini","Brescia","BS","studiomedicofilippini.it","1","Team (derma, endocrino, estetica)","Prof. Enrico Filippini (fondatore, 30+ anni)","Med. estetica viso/corpo, filler, dermatologia","MP+GEN","Multi-provider affermato + fondatore maturo (uscita generazionale), bacino Brescia ~200k","topdoctors.it/clinica/poliambulatorio-medico-chirurgico-prof-dott-enrico-filippini"],
 ["A","Lombardia (NO)","Lombardia","Poliambulatorio Finazzi","Bergamo","BG","poliambulatoriofinazzi.it","2 (medica + chirurgica)","Team specialisti","Anacleto Finazzi (fond.) / Glauco Finazzi (AD)","Med. estetica, laser chirurgia, chirurgia estetica","MP+GEN+PLAT","Realta familiare 2 sedi, laser dagli anni '90, dinamica di successione familiare","poliambulatoriofinazzi.it/la-nostra-storia"],
 ["B","Lombardia (NO)","Lombardia","Istituto Dermoclinico Vita Cutis","Milano","MI","dermoclinico.com","1+","Piu dermatologi","Prof. Antonino Di Pietro (fondatore)","Dermatologia estetica, anti-aging, med. rigenerativa, iniettabili","MP+GEN","Brand nazionale forte, multi-provider; caveat: premium, possibile resistenza alla cessione","dermoclinico.com/chi-siamo"],
 ["B","Lombardia (NO)","Lombardia","Studio Astolfi / CDM","Milano","MI","giorgioastolfi.it","1 (+ centro condiviso Settimo M.)","Founder + collaboratori","Dr. Giorgio Astolfi (1953, ~73 anni)","Filler, botox, PRP rigenerativa, laser CO2/vascolare, HIFU","GEN+CONV","Fondatore 73enne KOL iniettabili, core 'trattamenti ripetibili'; mono-medico da convertire","miodottore.it/giorgio-astolfi"],
 ["B","Lombardia (NO)","Lombardia","Laser Estetica / Polimedicalaser","Monza","MB","laserestetica.it","1 (+ Riccione)","Founder + team","Dr. Marcello Melandri (1979, 70+ anni; dal 1984)","Chirurgia laser, filler, blefaroplastica, rimozione tatuaggi","GEN+CONV","Segnale generazionale fortissimo ('primo centro laser d'Italia'); caveat: laser/chirurgia-heavy","laserestetica.it/about"],
 ["B","Lombardia (NO)","Lombardia","Clinica San Matteo (Poliamb. San Matteo)","Roncadelle (BS)","BS","poliambulatorisanmatteo.it","1","Staff multi-provider","Dir. San. Dr. Fabio G. Gritti","Med. + chirurgia estetica, dermatologia","MP","20+ anni, multi-provider affermato, hinterland Brescia","poliambulatorisanmatteo.it/la-clinica"],
 ["B","Lombardia (NO)","Lombardia","Centro Medico Ticinello","Pavia","PV","ticinello.it","2","Team ampio","n/d","Med. estetica (filler, botox, criolipolisi, peeling), dermatologia, chir. plastica","MP+CONV","Multi-provider 2 sedi, dept estetica strutturato, convertibile a hub","micuro.it/strutture/centro-medico-ticinello-di-pavia"],
 ["B","Lombardia (NO)","Lombardia","Centro Medico Ambrosiano","Milano","MI","centromedicoambrosiano.com","2","Multi-specialista","n/d","Botox, ac. ialuronico, rinofiller, radiofreq., med. rigenerativa, chir. plastica","MP+PLAT","Poliambulatorio privato multi-sede, servizi ripetibili + rigenerativa","centromedicoambrosiano.com/medicina-estetica"],
 ["B","Lombardia (NO)","Lombardia","PoliMed Poliambulatorio","Brescia","BS","polimed-brescia.it","1","Equipe (Cortellessa, Di Nonno, +)","n/d","Med. estetica, laser dermatologico, filler, epilazione, chir. plastica","MP","Poliambulatorio dedicato a estetica+laser, equipe strutturata","polimed-brescia.it/equipe"],
 ["C","Lombardia (NO)","Lombardia","Ambros Medica (Studio Sartori)","Milano","MI","ambrosmedica.it","1","Founder-led (+ staff)","Dr. Massimo Sartori (1991, ~30 anni)","Laser CO2/YAG/Alessandrite, filler, botox, fleboterapia","GEN+CONV","Founder maturo; caveat: molto laser/dermochirurgia, probabile mono-medico","hotfrog.it/company/1123239018147840"],
 ["C","Lombardia (NO)","Lombardia","Centro Medico Meldes","Milano","MI","meldes.it","1","n/d","Dr. Luca R. De Santis (fondatore)","Med. estetica 'naturale', anti-aging, laserterapia, filler/botox","CONV","Founder-led focus iniettabili/anti-aging; dati team/ricavi n/d","meldes.it"],
 ["C","Lombardia (NO)","Lombardia","Studio Medico Mapelli","Magenta (MI)","MI","studiomedicomapelli.it","1","3+ (Mapelli, Pedrazzoli, Tagliani)","Dr.ssa Elisabetta Mapelli (fond. 2012)","Botox, filler, biorivitalizzazione, terapia rigenerativa staminale","MP","Multi-provider in crescita; caveat: Magenta <100k, fondazione recente","studiomedicomapelli.it/blog/nuovo-poliambulatorio-mapelli"],
 ["C","Lombardia (NO)","Lombardia","Olosmed Poliambulatorio","Brescia","BS","olosmed.it","1","Multi-specialista","Dir. San. Dr. Emmanuele Ferrari","Med. estetica avanzata, laser, biorivitalizzazione, iniettabili","MP","Ambulatorio autorizzato Reg. Lombardia; caveat: giovane (2020)","medicinaesteticabrescia.net/chi-siamo"],
 ["C","Lombardia (NO)","Lombardia","Fortimed Italia","Azzano San Paolo (BG)","BG","fortimeditalia.it","1 (grande)","Molti specialisti","Dir. San. Dr. Gianluca Cacace","Med. estetica viso/corpo + laser (dept in poliambulatorio ampio)","MP+CONV","Multi-provider consolidato; caveat: estetica una branca tra molte","fortimeditalia.it/visite-esami/medicina-estetica-bergamo"],
 ["C","Lombardia (NO)","Lombardia","Modoetia Poliambulatorio Medico","Monza","MB","poliambulatoriomodoetia.it","1","Team (ex-ospedalieri)","Dir. San. Dr. Pietro Pizzi (dal 2019)","Filler, botulino, peeling, chir. plastica, nutrizione","MP","Multi-provider strutturato; caveat: giovane (2019), no angolo generazionale","poliambulatoriomodoetia.it"],
 ["C","Lombardia (NO)","Lombardia","Poliambulatorio Effetre","Cassano Magnago (VA)","VA","poliambulatorioeffetre.it","Multi (VA + Ticino CH)","n/d","n/d","Botox, filler, sclerosanti, biorivitalizzazione","PLAT","Piccola rete multi-sede cross-border, focus iniettabili; dati titolare n/d","poliambulatorioeffetre.it/medicina-estetica-varese"],
 ["C","Lombardia (NO)","Lombardia","Studio Dr. Carenzio (Poliambulatorio)","Pavia","PV","paviamedicinaestetica.it","1 (+ Mediperson)","Founder + chir. plastico + odontoiatra","Dr. Claudio Carenzio","Rinofiller, botox, filler labbra, dermochirurgia minore","GEN+CONV","Studio di famiglia trasformato in poliambulatorio (2018), angolo successione","paviamedicinaestetica.it"],
 ["C","Lombardia (NO)","Lombardia","Seven Suite","Seregno (MB)","MB","seven-suite.it","1","1 (chirurgo)","Dr. Ivan Arruda","Estetica rigenerativa, filler, botox, lifting non chir., laser","CONV","Focus estetica rigenerativa ripetibile; caveat: mono-medico, forte chirurgia","seven-suite.it/chi-siamo"],
 ["C","Lombardia (NO)","Lombardia","Derm Art Clinic","Varese","VA","dermartclinic.it","1","Founder + collaboratori","Dr.ssa Laura Balint (fondatrice)","Med. estetica, laser, botox, filler, fili, dermochirurgia","CONV","Clinica estetica dedicata; caveat: fondatrice giovane (no generazionale)","dermartclinic.it/drlaurabalint"],
 ["C","Lombardia (NO)","Lombardia","Polispecialistica Lariana (Dr.ssa Russo)","Como","CO","polispecialisticalariana.it","1","Poliamb.; estetica by Dr.ssa Russo","Dr.ssa Sara Russo","Botox, filler, HIFU, biostimolazione, mini-lifting non chir.","CONV","Presenza estetica a Como; caveat: professionista in affitto sale, non proprietaria","medicinaesteticasararusso.it"],
 ["C","Lombardia (NO)","Lombardia","CMP - Centro Medico Polispecialistico","Pavia","PV","centro-medico.it","1","Piu dermatologi","Prof. Nicola Zerbinati (dir. scientifico)","Dermatologia estetica, laser, iniettabili avanzati","MP","Reputazione/KOL fortissima, multi-provider; caveat: KOL accademico, probabile non-vendita","centro-medico.it/dottore-nicola-zerbinati"],

 # ---------------- PIEMONTE & LIGURIA (Nord-Ovest) ----------------
 ["A","Piemonte-Liguria (NO)","Piemonte","Poliambulatorio Medivela","Torino (Crocetta)","TO","medivela.com","1","Team multi-spec.","Nato 2011; Dir. San. Dr. Sergio Periotto","Filler, tossina, PRP, biostimolazione + poliambulatorio","MP+PLAT","Poliambulatorio strutturato/istituzionalizzato, non dipende da un nome: piattaforma-anchor Torino","medivela.com/medicina-estetica"],
 ["A","Piemonte-Liguria (NO)","Liguria","Clinica Visage","Genova (+ Ventimiglia)","GE","clinicavisage.it","2","Team","n/d (attiva 15+ anni)","Med. estetica, med. rigenerativa, longevity, iniettabili, laser","MP+PLAT","Multi-sede locale affermata, posizionamento rigenerativa/longevity moderno, baricentro Genova","clinicavisage.it"],
 ["B","Piemonte-Liguria (NO)","Piemonte","Studio Dr. Luigi Turco","Torino (+ Santena)","TO","medicinaesteticaturco.it","2","1 (titolare)","Dr. Luigi Turco (laurea 1980, ~70 anni)","Botulino, filler, laserterapia, carbossiterapia (+ odontoiatria)","GEN+CONV","Fondatore storico fascia matura, 2 sedi: uscita generazionale; caveat: mono-medico + mix odontoiatrico","medicinaesteticaturco.it/lo-studio"],
 ["B","Piemonte-Liguria (NO)","Piemonte","Studio Dermatologico Novarese","Novara","NO","studiodermatologiconovarese.it","1","2 (confermati)","Dr.ssa Simona Ferri + Dr. Christian Giani","Botox, filler, biorivit., PRP, laser + dermatologia","MP","Unico target con 2 medici confermati; mix derma+estetica+laser ricorrente; tuck-in Novara","studiodermatologiconovarese.it/chi-siamo"],
 ["B","Piemonte-Liguria (NO)","Piemonte","Dott. Piercarlo Masolini","Alessandria (+ Tortona/Asti/Casale)","AL","piercarlomasolini.it","Multi (4 citta)","1 (titolare)","Dr. Piercarlo Masolini (1959, ~67 anni)","Filler, botulino, Plexr, med. rigenerativa, laser (+ odontoiatria)","GEN+PLAT+CONV","Titolare maturo + rete multi-citta + rigenerativa; caveat: mix odontoiatrico da carve-out","piercarlomasolini.it/medicina-estetica"],
 ["C","Piemonte-Liguria (NO)","Piemonte","Poliambulatorio MeD Italia","Torino (Crocetta)","TO","med-italia.it","1","Staff multi-spec.","n/d - SRL poliambulatorio","Med. estetica + diagnostica specialistica","MP+CONV","Poliambulatorio indipendente zona pregio, convertibile a hub; dati n/d","med-italia.it/medicina-estetica"],
 ["C","Piemonte-Liguria (NO)","Piemonte","Dott.ssa Laura Ferrero","Torino","TO","lauraferrero.it","1","1 (titolare)","Dr.ssa Laura Ferrero (laurea 1994, studio 2001)","Filler, botulino, biorivit., radiofreq., laser + nutrizione","CONV","Studio affermato 25 anni, focus rigenerativa/iniettabili; caveat: mono, mid-career","lauraferrero.it/medicina-estetica"],
 ["C","Piemonte-Liguria (NO)","Piemonte","Prof. Franco Buttafarro","Torino","TO","skinguru.it","1","1 (titolare)","Prof. Franco Buttafarro (laurea 1976, ~73)","Botox, filler, laser + dermatologia + chir. estetica/tricologia","GEN+CONV","Fondatore storico molto maturo; caveat: forte componente chirurgica/dermatologica, mono","skinguru.it"],
 ["C","Piemonte-Liguria (NO)","Piemonte","Polispecialistico Leonardo","Novara","NO","polispecialisticoleonardo.it","1","Staff multi-spec.","n/d - SRL","Med. estetica + dermatologia + odontoiatria + specialistica","MP+CONV","Struttura multi-provider avviata; caveat: estetica una linea fra molte","polispecialisticoleonardo.it/medicina-chirurgia-estetica-novara"],
 ["C","Piemonte-Liguria (NO)","Liguria","Studio Dott. Christian Molinelli","Genova","GE","dottmolinelligenova.com","1","1 (titolare)","Dr. Christian Molinelli","HIFU, filler, botox, MesoJet, epilazione laser, ozonoterapia","CONV","Device+iniettabili ripetibili; caveat: mono-provider, dimensioni piccole","dottmolinelligenova.com"],
 ["C","Piemonte-Liguria (NO)","Liguria","Dott. Giuseppe Manzo","Genova","GE","medicinaesteticagenova.com","1","1 (titolare)","Dr. Giuseppe Manzo","Botox, filler, biostimolazione + chir. plastica/liposcultura","CONV","Iniettabili; caveat: orientamento chirurgico, mono","medicinaesteticagenova.com"],
 ["C","Piemonte-Liguria (NO)","Liguria","Dott.ssa Lucrezia Sugliano","Genova / Sanremo / Milano","GE","lucreziasugliano.it","3 (presso centri terzi)","1 (titolare)","Dr.ssa Lucrezia Sugliano","Filler, neuromodulazione, biostimolazione (full-face)","CONV","Focus puro iniettabili; caveat: opera dentro centri terzi (non asset-clinica autonomo)","lucreziasugliano.it/chi-sono"],
 ["C","Piemonte-Liguria (NO)","Liguria","Centro Medico Minerva","Savona","SV","centromedicominerva.it","1 (550 mq)","Team multi-spec.","Med. estetica by Dr.ssa Simona Grosso (chir. plastico)","Filler, botulino, laser, radiofreq. + poliambulatorio","MP+CONV","Poliambulatorio ampio con sala chir. ambulatoriale; caveat: struttura recente","centromedicominerva.it/servizio-medicina-estetica-43"],
 ["C","Piemonte-Liguria (NO)","Liguria","Levante Laser Studio","La Spezia","SP","levantelaserstudio.com","1","Team; Dr.ssa R. Cestari (derma)","n/d","Filler, botox, biorivit., laser (epilazione/tatuaggi/macchie), derma","MP+CONV","Struttura autorizzata, mix laser+iniettabili+derma ripetibile; dati n/d","levantelaserstudio.com"],
 ["C","Piemonte-Liguria (NO)","Liguria","Clinica Med","La Spezia","SP","clinicamed.it","1","Team","n/d","Filler, botox, carbossiterapia, needling, peeling, electro-sculpting","CONV","Focus med. estetica non chirurgica, trattamenti ripetibili; dimensione da verificare","clinicamed.it/medicina-estetica-la-spezia"],

 # ---------------- NORD-EST ----------------
 ["A","Nord-Est","Veneto","Medical Laser Clinic","San Martino B.A. (Verona)","VR","medicallaserclinic.it","1","Multi (equipe)","Dr. Stefano Anderluzzi (socio fond.) + Dr. Francesco Colla","Iniettabili, laserterapia, dermatologia, chir. vascolare, nutrizione","MP","Multi-provider indipendente, 2 soci medici, mix laser+iniettabili+derma, bacino Verona","medicallaserclinic.it/il-centro"],
 ["A","Nord-Est","Veneto","AES Clinic (Perla Medicina Srl)","Padova","PD","aesclinic.it","1","Multi (equipe)","Dott.ssa MianMian Wang (laurea 2003)","Iniettabili, laser, anti-aging, no chirurgia pesante","MP","Profilo piu vicino all'ICP puro: indipendente, multi-provider, non-chirurgico, centro Padova","aesclinic.it/medici-medicina-estetica-padova"],
 ["A","Nord-Est","Emilia-Romagna","Centro San Prospero - Med. e Chir. Estetica","Bologna","BO","medicinaesteticasanprospero.it","1 (sala operatoria interna)","Multi (5 specialita)","Tra i fondatori Antonio Gotti (10+ anni)","Radiofreq., biorivit., botulino, filler + chirurgia","MP","Multi-provider indipendente maturo (chir. plastico, medico estetico, derma, flebologo, dietista), hub Bologna","medicinaesteticasanprospero.it/medicina-chirurgia-estetica-bologna"],
 ["A","Nord-Est","Emilia-Romagna","Studio Dott.ssa Clelia Barini","Modena (Formigine)","MO","medicinaesteticabarini.it","2 (Modena, La Spezia)","1 (fondatrice)","Dott.ssa Clelia Barini (dal 1994, ~30 anni)","Filler, botulino, fili, peeling, laser frazionato, Thermage, radiofreq.","GEN+PLAT","Segnale generazionale forte (30 anni, consiglio SIES, docente UniMoRe), multi-sede + energy-based","medicinaesteticabarini.it/Gli-studi"],
 ["A","Nord-Est","Emilia-Romagna","Studio Dott.ssa Paola Molinari","Modena","MO","paolamolinari.it","1","1 + staff","Dott.ssa Paola Molinari (studio dal 1987, 30+ anni)","Filler, botox, trattamenti viso","GEN","Candidata succession di prim'ordine: 30+ anni, Teoxane Training Center, docente; brand trasferibile","paolamolinari.it/studio-medicina-estetica-molinari"],
 ["B","Nord-Est","Veneto","Studio Dr. Franco Barbazza","Padova","PD","francobarbazza.it","1","1 (titolare)","Dr. Franco Barbazza (laurea 1988, ex-primario univ.)","Filler, botulino, biorivit., fili, sclero, mesoterapia, radiofreq. + chirurgia","GEN+CONV","Fondatore maturo ex-primario (uscita generazionale), citta 100k+; caveat: mono + chirurgia","francobarbazza.it/studio-medico-padova"],
 ["B","Nord-Est","Friuli-VG","Studio Dott.ssa Erica Antonini","Trieste (+ Conegliano)","TS","ericaantoninitrieste.it","2 (c/o Fisiomed)","1 (titolare)","Dott.ssa Erica Antonini (laurea 1994, 30+ anni)","Laser (CO2, Fraxel, ND:YAG), med. estetica, filler, botox, biorivit.","GEN+CONV","Fondatrice matura forte su laser/derma; caveat: opera dentro poliambulatorio Fisiomed","ericaantoninitrieste.it/trattamenti-laser"],
 ["B","Nord-Est","Veneto","AD Aesthetics","Verona (+ Milano)","VR","adaesthetics.it","2","Multi (team)","Dr. Alessandro Dall'Antonia (chir. plastico, dir. san.)","Med. estetica, iniettabili, med. rigenerativa, chir. mini-invasiva","MP+PLAT","Multi-provider + rigenerativa, gia multi-citta; caveat: sbilanciata su chirurgia","adaesthetics.it/ad-aesthetics/il-nostro-team"],
 ["B","Nord-Est","Veneto","Medical Art - Soft Surgery Clinic","Vicenza (+ Treviso)","VI","medicalart.it","Multi","Multi","Dir. San. Dr. Domenico Miccolis (27 anni exp.)","Med. estetica, dermatologia, chirurgia 'soft', iniettabili","MP+PLAT","Multi-provider multi-sede; caveat: parte del brand EA Aesthetics (mini-catena), sbilancio chirurgia","medicalart.it"],
 ["B","Nord-Est","Emilia-Romagna","Centro Medicina Estetica Elisa Benati","Carpi / Modena","MO","elisabenati.it","2 (Modena, Carpi)","1 (titolare)","Dott.ssa Elisa Benati","Filler, botox, biorivit., peeling, laser","PLAT+CONV","Indipendente multi-sede locale; iniettabili ripetibili","elisabenati.it/medicina-estetica"],
 ["C","Nord-Est","Veneto","Studio Dr. Antonello Goglia","Verona (Borgo Trento)","VR","medicinaestetica-goglia.com","1","1 (n/d)","Dr. Antonello Goglia","Medicina estetica iniettabili","CONV","Indipendente, citta 100k+; dati team limitati","medicinaestetica-goglia.com"],
 ["C","Nord-Est","Veneto","Studio Medico Santa Marina","Venezia","VE","studiomedicosantamarina.it","1","n/d","n/d","Dermatologia estetica, laserterapia, med. estetica, anti-aging","CONV","Indipendente citta 100k+; dati team/fondatore limitati","studiomedicosantamarina.it/dermatologia-estetica"],
 ["C","Nord-Est","Emilia-Romagna","Studio Medico Paris","Bologna","BO","vivianaparis.it","1","1 (titolare)","Dott.ssa Viviana Paris","Med. estetica non-chirurgica, tecnologie viso/corpo","CONV","Indipendente, citta 100k+; mono","vivianaparis.it"],
 ["C","Nord-Est","Emilia-Romagna","Giorgini Clinique","Bologna","BO","giorginiclinique.it","1","1 (titolare)","Dott.ssa Beatrice Giorgini","Med. estetica non invasiva viso, iniettabili, tecnologie","CONV","Indipendente, citta 100k+; mono","giorginiclinique.it/medicina-estetica-bologna"],
 ["C","Nord-Est","Emilia-Romagna","Studio Dott.ssa Anna Cristina Mazzini","Modena","MO","annacristinamazzinimedicinaestetica.it","1","1 (titolare)","Dott.ssa Anna Cristina Mazzini","Filler, laserterapia, trattamenti viso/corpo","CONV","Indipendente, citta 100k+; mono","annacristinamazzinimedicinaestetica.it"],
 ["C","Nord-Est","Emilia-Romagna","Medicina Estetica Papalia","Reggio Emilia","RE","medicinaesteticapapalia.it","1","n/d","Papalia","Med. estetica viso/corpo, tecnologie elettromedicali avanzate","CONV","Indipendente citta 100k+; dati fondatore/team limitati","medicinaesteticapapalia.it/servizi"],
 ["C","Nord-Est","Emilia-Romagna","Studio Dott.ssa Roberta Momi","Reggio Emilia","RE","robertamomi.it","1","1 (titolare)","Dott.ssa Roberta Momi (~20 anni exp.)","Med. estetica viso/corpo, filler, botulino","CONV","Professionista affermata, citta 100k+; mono","robertamomi.it"],
 ["C","Nord-Est","Emilia-Romagna","Medicina Estetica Parma (Studio Kachalkina)","Parma","PR","medicinaesteticaparma.it","1","1 (titolare)","Dott.ssa Anna Kachalkina","Filler, botox, biorivit., laser epilazione, blefaroplastica laser","CONV","Indipendente, citta 100k+; mono","medicinaesteticaparma.it"],
 ["C","Nord-Est","Emilia-Romagna","EnjoyDerma","Ferrara","FE","enjoyderma.it","1","Multi (polispecialistico)","Dir. San. Dott.ssa Valeria Scuderi","Dermatologia, med. estetica, laser (incl. Aviclear), nutrizione","MP","Multi-provider, laser avanzato; caveat: struttura molto giovane (~1 anno), ricavi probab. sotto ICP","enjoyderma.it/medicina-estetica"],
 ["C","Nord-Est","Friuli-VG","Studio Dott.ssa Euridice Fabris","Trieste","TS","fabriseuridice.com","1","1 (titolare)","Dott.ssa Euridice Fabris (laurea 1997)","Filler, botulino, fili, peeling, radiofreq., laser vascolare/epilazione","CONV","Indipendente, citta 100k+; mono","fabriseuridice.com"],
 ["C","Nord-Est","Friuli-VG","Studio Dott. Marino Trivisani","Udine","UD","medicinaesteticadottmarinotrivisani.it","1","1 (titolare)","Dr. Marino Trivisani","Filler, botox, laser diodo, blefaroplastica non chir., trattamenti pelle","CONV","Indipendente, citta 100k+; mono","medicinaesteticadottmarinotrivisani.it/chi-siamo"],
 ["C","Nord-Est","Trentino-AA","Poliambulatorio MeSS","Trento","TN","mess-srl.it","1","Multi-provider","n/d","Laser, filler, botulino","MP+CONV","Poliambulatorio multi-provider; dati fondatore n/d","mess-srl.it"],
 ["C","Nord-Est","Trentino-AA","Odontomedica - Studio Detassis","Trento","TN","studiodetassis.com","1","Team","Dr. Danilo Detassis (titolare dal 1991)","Filler, PRP, lipofilling (add-on a odontoiatria)","GEN+CONV","Fondatore maturo (succession); caveat: odontoiatria-primaria, estetica add-on (fit marginale)","studiodetassis.com"],

 # ---------------- ROMA & CENTRO ----------------
 ["A","Roma-Centro","Lazio","Studio Medico Estetico Monti Parioli","Roma","RM","mariangelaercoli.com","1 (Parioli)","Piu medici + staff","Dott.ssa Mariangela Ercoli (studio prof.; 'Mariangela Ercoli Srl' risulta IN LIQUIDAZIONE)","Med. estetica, filler/botox, chirurgia plastica, benessere","MP+GEN","Fondatrice storica 40+ anni, Parioli premium; MA assetto societario da chiarire (SRL in liquidazione + doppia identita web mariangelaercoli.com / montipariolimedical.it). Attenzione omonimia 'Monti Parioli Medical' (Del Parco)","guidaestetica.it/centri/studio-medico-estetico-monti-parioli"],
 ["A","Roma-Centro","Lazio","Studio Saccomanno / SabaMedica Longevity HUB","Roma","RM","studiosaccomanno.it","2 (Piramide + San Saba)","Team multidisciplinare","Dott.ssa Karin Saccomanno (Pres. AISI)","Med. estetica, longevity/prevenzione, chirurgia specialistica, nutrizione","MP+PLAT","Indipendente per vocazione (titolare presiede Assoc. Sanita Indipendente), 2 sedi, brand premium","studiosaccomanno.it/medicina-estetica"],
 ["A","Roma-Centro","Lazio","Centro Medico Estetico Raffaella Casilli","Roma (+ Anzio)","RM","raffaellacasilli.com","2","1 (libera prof.)","Dott.ssa Raffaella Casilli (30 anni exp.)","Filler, botox, laser frazionale, luce pulsata, radiofreq., needling, carbossiterapia","GEN+PLAT","30 anni di esperienza (generazionale), 2 sedi, mix iniettabili + energy-based ripetibili","raffaellacasilli.com/chi-sono-2"],
 ["A","Roma-Centro","Toscana","Clinica Gramsci - Med. e Chirurgia Estetica","Firenze / Prato","FI","clinicagramsci.it","2 (Firenze + Prato)","3+ (Ghezzi, Quercioli, Falconi + team)","Dott.ssa Serena Ghezzi + Dr. Fabio Quercioli","Med. estetica, med. rigenerativa, laserterapia, chir. plastica","MP+PLAT","Gia rete 2 sedi con divisioni distinte estetica/rigenerativa/laser, entita P.IVA dedicata; piattaforma toscana","ufficiocamerale.it (P.IVA 07363720488)"],
 ["A","Roma-Centro","Abruzzo/Umbria","JFK Medical Beauty","Pescara / Perugia","PE","jfkmedicalbeauty.it","2 (Pescara + Perugia)","Team (es. C. Carusi, C. Carrozzo)","JFK Medical Beauty","Botox, filler, skinbooster, PRP, fili, laser Nd:YAG/CO2","MP+PLAT","Rete interregionale gia impostata, core iniettabili ripetibili: mini-piattaforma Abruzzo+Umbria","jfkmedicalbeauty.it/medicina-estetica-pescara"],
 ["B","Roma-Centro","Lazio","Tila Institute","Roma","RM","tilainstitute.com","1 (Roma Nord)","Dir. + staff","Dasantila Myftari (fond. 2008); dir. san. M. Marcellino","Med. estetica viso/corpo, dermatologia, chirurgia estetica","MP+CONV","Fondata 2008, brand riconoscibile, convertibile multi-provider; zona Roma Nord benestante","tilainstitute.com/Il-Centro"],
 ["B","Roma-Centro","Lazio","Aventino Medical Group","Roma","RM","aventinomedicalgroup.com","1 (Aventino)","Molti specialisti + sala operatoria","Gruppo internazionale (poliambulatorio)","Dermatologia, med. estetica, mesoterapia, chirurgia estetica","MP+CONV","Struttura ampia con reparto dermato-estetico: possibile carve-out/piattaforma","aventinomedicalgroup.com/en/aesthetic-medicine"],
 ["B","Roma-Centro","Lazio","Clinic Cinecitta","Roma","RM","cliniccinecitta.it","1 (Cinecitta)","Piu medici","Clinic Cinecitta","Med. + chir. estetica, filler, botox, biostimolazioni","MP","Bacino Roma Sud-Est, mix trattamenti ripetibili, struttura clinica","cliniccinecitta.it"],
 ["B","Roma-Centro","Lazio","Domimedica","Roma","RM","domimedica.it","1","Piu medici","Domimedica (chir. plastica + med. estetica)","Laser CO2, filler, botox, chirurgia plastica","MP","Centro affermato, presenza energy-based (laser CO2) + iniettabili","domimedica.it/medicina-estetica-roma"],
 ["B","Roma-Centro","Lazio","Clinica Parioli","Roma","RM","clinicaparioli.it","1 (Parioli)","Piu medici","Clinica Parioli","Med. estetica + laserterapia (adiposita, cicatrici, ringiovanimento)","MP+CONV","Clinica strutturata zona premium, energy-based, convertibile","clinicaparioli.it/ambulatori-e-chirurgie/medicina-estetica-e-laser-terapia-a-roma"],
 ["B","Roma-Centro","Lazio","Polistudio Le Rose","Roma","RM","polistudiolerose.com","1","Piu specialisti","Polistudio Le Rose","Filler, botox, radiofreq., med. estetica viso non invasiva","MP+CONV","Poliambulatorio con reparto estetico dedicato, convertibile multi-provider","polistudiolerose.com/servizi/medicina-estetica"],
 ["B","Roma-Centro","Toscana","Centro Medico Laser Firenze","Firenze","FI","centromedicolaser.it","2 (Firenze + Scandicci)","Dir. sanitario + medico estetico","Centro Medico Laser","Med. estetica, laserterapia, dermatologia, dep. laser","MP+PLAT","Realta affermata da molti anni, 2 sedi, core energy-based ripetibile","centromedicolaser.it/medicina-estetica-firenze"],
 ["B","Roma-Centro","Toscana","Istituto Medlight","Firenze","FI","medlight.it","1","Team medici estetici + chir. plastici + dermatologi","Istituto Medlight","Med. estetica, laserterapia, chirurgia estetica, elettromedicale","MP","Team multi-provider consolidato, parco laser/elettromedicale avanzato","medlight.it"],
 ["B","Roma-Centro","Toscana","Clinica Ireos","Firenze","FI","clinicaireos.com","1","Team qualificato","Clinica Ireos","Med. + chirurgia estetica, iniettabili, tecnologie","MP+CONV","Clinica affermata a Firenze, mix estetica/energy-based, convertibile","clinicaireos.com"],
 ["B","Roma-Centro","Toscana","Istituto Medico Toscano (IMT / Alma Vita Srl)","Prato","PO","istitutomedicotoscano.it","1 (3.000 mq)","~50 dipendenti, piu specialisti","Alma Vita S.r.l.","Med. estetica e rigenerativa, PRP, chirurgia, odontoiatria","MP+CONV","Unico centro Toscana abilitato PRP; reparto estetico dentro poliambulatorio (carve-out/piattaforma)","miodottore.it/strutture/alma-vita-s-r-l"],
 ["B","Roma-Centro","Umbria","Dott. Francesco Bachiorri","Perugia / Citta di Castello","PG","bachiorri.com","2-3 (Perugia, C. di Castello, Bologna)","1 (chir./medico estetico)","Dr. Francesco Bachiorri (attivo dal 1997)","Med. estetica (filler, botox) + chirurgia estetica","GEN+PLAT+CONV","Profilo maturo/generazionale, gia rete multi-citta; caveat: mono + chirurgia","bachiorri.com"],
 ["C","Roma-Centro","Lazio","Studio Lydia Igoumenaki (Roma Med. Estetica)","Roma","RM","romamedicinaestetica.it","1","1 (+ collab.)","Dott.ssa Lydia Igoumenaki","Iniettabili viso/labbra, filler, botox","CONV","Alta notorieta/volumi (373 recensioni), focus iniettabili ripetibili; caveat: mono-provider","romamedicinaestetica.it"],
 ["C","Roma-Centro","Lazio","Studio Borreo","Roma","RM","studioborreo.it","1 (Parioli)","n/d","Studio Borreo","Med. estetica viso/corpo, iniettabili","CONV","Zona Parioli premium, studio indipendente focalizzato; dati n/d","studioborreo.it/medicina-estetica"],
 ["C","Roma-Centro","Lazio","Ellesmile","Roma","RM","ellesmile.it","1 (Trastevere)","Specialisti","Ellesmile","Filler, botox, biostimolazione, med. estetica","CONV","Studio autorizzato zona centrale Trastevere, iniettabili ripetibili","ellesmile.it/medicina-estetica-roma-trastevere"],
]

# Candidati minori / da valutare in fase due (fit debole) - foglio a parte
minor = [
 ["Roma-Centro","Marche","Dott.ssa Veronica Consales","Ancona","AN","veronicaconsales.it","Plexr, radiofreq., luce pulsata, laser frazionale, botox, filler","Ampio energy-based ma studio mono-medico; migliore opzione 'pura' ad Ancona"],
 ["Roma-Centro","Toscana","Dott. Michele D'Anteo","Livorno","LI","drdanteo.it","Dermatologia estetica","Dermatologia estetica affermata; mono"],
 ["Roma-Centro","Abruzzo","Studio Regina","Pescara","PE","studiodentisticoregina.it","Piano dedicato a estetica","Base e studio odontoiatrico -> fit debole"],
 ["Roma-Centro","Toscana","INLAB","Livorno","LI","in-lab.it","Med. estetica in centro analisi","Med. estetica dentro poliambulatorio analisi -> fit debole"],
]

# Esclusi (consolidatori/catene/PE) - foglio a parte
excluded = [
 ["Ariel Medical","Catena nazionale (Milano/Torino/Bologna/Genova/Firenze), fond. 2020, importa tecnologia beauty israeliana","Consolidatore/concorrente, non target","arielmedical.it/chi-siamo"],
 ["SkinMedic","Brand/franchising nazionale di estetica avanzata (store locator)","Catena/franchising, fuori ICP","skinmedic.it"],
 ["ZucchiSkin Center / Vein & Derma Clinic","Gruppo San Donato (capitale PE/ospedaliero)","Gia consolidato da gruppo, fuori ICP","-"],
 ["Synlab CAM (Monza)","Catena Synlab","Catena internazionale, fuori ICP","-"],
 ["CDI Milano","Grande gruppo diagnostico","Fuori scala/ICP","-"],
 ["theClinic / Nika Clinic (Milano)","Realta branded milanesi di grande scala","Da riclassificare solo per upper-market","-"],
]

wb = openpyxl.Workbook()

# ---- Palette ----
NAVY = "3E1E2E"; WINE = "6E2A44"; ROSE = "F0EBE6"; GOLD = "B08A6E"
white = Font(color="FFFFFF", bold=True, size=11)
hdr_fill = PatternFill("solid", fgColor=WINE)
tierfill = {"A": PatternFill("solid", fgColor="C6E7C9"),
            "B": PatternFill("solid", fgColor="FDF3C9"),
            "C": PatternFill("solid", fgColor="F3E1E1")}
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(vertical="center", horizontal="center")

# ================= SHEET 1: Longlist =================
ws = wb.active
ws.title = "Longlist"
headers = ["Tier","Cluster","Regione","Clinica","Citta","Prov","Sito web","N. sedi",
           "N. medici","Titolare/Fondatore","Servizi principali","Angolo",
           "Segnali di fit / caveat","Fonte"]
ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(1, c); cell.font = white; cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = border
for r in rows:
    ws.append(r)
# style body
for ri in range(2, ws.max_row+1):
    tier = ws.cell(ri,1).value
    for ci in range(1, len(headers)+1):
        cell = ws.cell(ri, ci); cell.alignment = wrap; cell.border = border
    ws.cell(ri,1).fill = tierfill.get(tier, PatternFill()); ws.cell(ri,1).alignment = center
    ws.cell(ri,1).font = Font(bold=True)
    ws.cell(ri,6).alignment = center
widths = [6,20,16,34,22,6,30,14,26,34,44,14,52,40]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:N{ws.max_row}"
ws.row_dimensions[1].height = 30

# ================= SHEET 2: Sintesi =================
ws2 = wb.create_sheet("Sintesi")
def h(row, text, size=14, color=WINE):
    c = ws2.cell(row,1, text); c.font = Font(bold=True, size=size, color=color); return c
h(1, "PROGETTO VENERE - Longlist target medicina estetica (Italia)", 16, NAVY)
ws2.cell(2,1,"Buy-and-build: consolidamento della medicina estetica IT. Deliverable di origination - 26/07/2026")
ws2.cell(2,1).font = Font(italic=True, color="666666")

# counts
from collections import Counter
clusters = Counter(r[1] for r in rows)
tiers = Counter(r[0] for r in rows)
h(4,"Copertura", 13)
ws2.cell(5,1,"Totale target in longlist"); ws2.cell(5,2, len(rows)).font = Font(bold=True)
row = 6
for cl,n in clusters.items():
    ws2.cell(row,1, "  "+cl); ws2.cell(row,2, n); row+=1
h(row+0,"Tiering", 13);
ws2.cell(row+1,1,"  Tier A - priorita di approccio (fit ICP piu forte)"); ws2.cell(row+1,2, tiers["A"])
ws2.cell(row+2,1,"  Tier B - solido con caveat"); ws2.cell(row+2,2, tiers["B"])
ws2.cell(row+3,1,"  Tier C - da validare / fit minore"); ws2.cell(row+3,2, tiers["C"])

# Top picks
tp_row = row+5
h(tp_row,"TOP 12 (Tier A) - priorita di approccio", 13)
tp_row += 1
ws2.cell(tp_row,1,"Clinica").font=Font(bold=True); ws2.cell(tp_row,2,"Citta").font=Font(bold=True)
ws2.cell(tp_row,3,"Angolo").font=Font(bold=True); ws2.cell(tp_row,4,"Perche").font=Font(bold=True)
tp_row+=1
for r in [x for x in rows if x[0]=="A"]:
    ws2.cell(tp_row,1, r[3]); ws2.cell(tp_row,2, f"{r[4]} ({r[5]})")
    ws2.cell(tp_row,3, r[11]); ws2.cell(tp_row,4, r[12])
    for ci in range(1,5): ws2.cell(tp_row,ci).alignment = wrap
    tp_row+=1
for i,w in enumerate([34,26,14,80],1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# legend
lg = tp_row+1
h(lg,"Legenda 'Angolo'", 12)
for i,(k,v) in enumerate([
    ("MP","Multi-provider gia strutturato (2+ medici / poliambulatorio)"),
    ("GEN","Uscita generazionale (fondatore maturo / 30+ anni di attivita)"),
    ("PLAT","Piattaforma / gia multi-sede (bolt-on, roll-up)"),
    ("CONV","Mono-medico o struttura da convertire a multi-provider")]):
    ws2.cell(lg+1+i,1, k).font=Font(bold=True); ws2.cell(lg+1+i,2, v)

# ================= SHEET 3: ICP & Metodo =================
ws3 = wb.create_sheet("ICP e Metodo")
lines = [
 ("Criteri del target ideale (dal deck Venere, pag. 9)", 13, WINE),
 ("- Ricavi EUR 0,8-2,5M, EBITDA margin 20%+ (verificato in DD con normalizzazione ricavi)", 11, "000000"),
 ("- Citta con bacino 100k+ abitanti; priorita cluster Nord-Ovest, Nord-Est, Roma", 11, "000000"),
 ("- Base pazienti attiva 1.500+, prevalenza iniettabili/trattamenti ripetibili", 11, "000000"),
 ("- 2+ medici operativi, o struttura convertibile a multi-provider entro 12 mesi", 11, "000000"),
 ("- Titolare disponibile a permanenza clinica 24-36 mesi con earn-out e rollover", 11, "000000"),
 ("- Autorizzazioni sanitarie in regola e sede idonea all'ampliamento", 11, "000000"),
 ("", 11, "000000"),
 ("Struttura tipo dell'operazione", 13, WINE),
 ("- Prezzo d'ingresso 3-4x EBITDA normalizzato", 11, "000000"),
 ("- 60-70% cash al closing, 15-20% earn-out a 24 mesi, 15-25% rollover equity", 11, "000000"),
 ("", 11, "000000"),
 ("Metodo di ricerca e limiti", 13, WINE),
 ("- Longlist costruita via ricerca web su 4 cluster (Lombardia; Piemonte-Liguria; Nord-Est; Roma-Centro).", 11, "000000"),
 ("- Ogni clinica e reale e con fonte URL. Le realta franchising/PE/catene sono state escluse (vedi foglio Esclusi).", 11, "000000"),
 ("- IMPORTANTE: ricavi ed EBITDA NON sono pubblici per questi studi/SRL privati -> per tutti valgono 'n/d - stima'.", 11, "B00000"),
 ("- Prossimo step: pull camerale (visure/bilanci Registro Imprese) sulle SRL identificate per stimare fatturato vs soglia ICP,", 11, "000000"),
 ("  ed enrichment dei contatti dei titolari (es. via Lusha) sui Tier A prima dell'approccio.", 11, "000000"),
 ("- Le autorizzazioni sanitarie regionali vanno confermate puntualmente in due diligence.", 11, "000000"),
]
for i,(t,s,c) in enumerate(lines,1):
    cell = ws3.cell(i,1,t); cell.font = Font(bold=(s>11), size=s, color=c); cell.alignment = wrap
ws3.column_dimensions["A"].width = 120

# ================= SHEET 4: Candidati minori =================
ws4 = wb.create_sheet("Candidati minori")
mh = ["Cluster","Regione","Nome","Citta","Prov","Sito","Servizi","Note (fit debole)"]
ws4.append(mh)
for c in range(1,len(mh)+1):
    ws4.cell(1,c).font=white; ws4.cell(1,c).fill=hdr_fill; ws4.cell(1,c).border=border
for m in minor: ws4.append(m)
for ri in range(2, ws4.max_row+1):
    for ci in range(1,len(mh)+1):
        ws4.cell(ri,ci).alignment=wrap; ws4.cell(ri,ci).border=border
for i,w in enumerate([18,14,30,16,6,28,40,50],1):
    ws4.column_dimensions[get_column_letter(i)].width=w

# ================= SHEET 5: Esclusi =================
ws5 = wb.create_sheet("Esclusi")
eh = ["Nome","Cosa e","Motivo esclusione","Fonte"]
ws5.append(eh)
for c in range(1,len(eh)+1):
    ws5.cell(1,c).font=white; ws5.cell(1,c).fill=hdr_fill; ws5.cell(1,c).border=border
for e in excluded: ws5.append(e)
for ri in range(2, ws5.max_row+1):
    for ci in range(1,len(eh)+1):
        ws5.cell(ri,ci).alignment=wrap; ws5.cell(ri,ci).border=border
for i,w in enumerate([34,54,44,28],1):
    ws5.column_dimensions[get_column_letter(i)].width=w

# ================= SHEET 6: Tier A - Dossier (camerale + contatti) =================
# Colonne: Clinica, Cluster, Ragione sociale / Forma, P.IVA, Fatturato (anno/fonte),
#          Verdetto ICP, Referente, Telefono, Email, PEC, LinkedIn, Indirizzo, Flag/caveat
dossier = [
 ["Studio Medico Estetico Monti Parioli","Roma","Studio prof. / ditta indiv.; 'Mariangela Ercoli Srl' IN LIQUIDAZIONE","n/d","n/d (non depositato)","n/d","Dott.ssa Mariangela Ercoli","06 87089918 / 351 4676449 (sito attuale); 06 3226547 (reg.)","info@montipariolimedical.it","n/d","n/d","Via dei Monti Parioli 36, 00197 Roma","Assetto societario da chiarire: SRL in liquidazione + doppia identita web. Omonimia con 'Monti Parioli Medical' (sorelle Del Parco)"],
 ["Studio Saccomanno / SabaMedica","Roma","SAN SABA MEDICA S.R.L. (ATECO 86.22.09)","n/d","fascia EUR 0,3-0,6M (2023) - informazione-aziende.it","Probabile SOTTO","Dott.ssa Karin Saccomanno","06 5757308 / 349 7391039","st.saccomanno@gmail.com","n/d","n/d","Viale Piramide Cestia 1, 00153 Roma","Possibile upside se sommato il fatturato su P.IVA personale della dottoressa (non consolidato)"],
 ["Centro Medico Estetico Raffaella Casilli","Roma","Studio prof. / ditta indiv. (nessuna SRL)","n/d","n/d (non depositato)","Probabile SOTTO","Dott.ssa Raffaella Casilli","06 3220030 / 366 1914536","info@raffaellacasilli.com","n/d","linkedin.com/in/dottoressa-raffaella-casilli","Viale Belle Arti 7, 00196 Roma (+ Anzio)","Struttura mono-professionale su 2 studi (appuntamento)"],
 ["Clinica Gramsci","Firenze/Prato","CLINICA GRAMSCI - CHIRURGIA PLASTICA E MEDICINA ESTETICA S.R.L.","07363720488","n/d (SRL con bilancio depositato - non accessibile via web)","n/d - PRIORITARIA visura","Dott.ssa Serena Ghezzi / Dr. Fabio Quercioli","055 8358297","info@clinicagramsci.it","n/d (esiste, non esposta)","linkedin.com/in/dott-serenaghezzi ; linkedin.com/in/fabio-quercioli-9319b042","Viale Gramsci 63/65, Firenze (+ V.le Marconi 50/7, Prato)","Unica vera SRL strutturata con bilancio tra i Tier A Roma-Centro: prioritaria per visura camerale a pagamento (i numeri esistono)"],
 ["JFK Medical Beauty","Pescara/Perugia","Non identificata (nessuna 'JFK Srl' confermata)","n/d","n/d","n/d","n/d (Dr. Carlo Carusi, chirurgo)","085 2010818","info@jfkmedicalbeauty.it","n/d","n/d","Viale J.F. Kennedy 167, 65123 Pescara; Perugia n/d","Ragione sociale mancante; discordanza indirizzo Pescara (Kennedy 167 vs 'via Bologna'); NON usare 'JMJ Pescara Srls' (non provato il legame)"],
 ["Poliambulatorio Filippini","Brescia","Studio / ditta indiv. (probabile)","03017310172","n/d (non depositato)","n/d","Prof. Enrico Filippini","030 2807547 / 335 5850800 / WA 389 2681259","info@studiomedicofilippini.it","n/d","linkedin.com/company/poliambulatoriofilippini (aziendale)","Via San Francesco d'Assisi 3, 25122 Brescia","OMONIMIA: 'Poliambulatorio Filippini S.R.L.' P.IVA 02584010124 e' a Tradate (VA) - entita DIVERSA, non confondere"],
 ["Poliambulatorio Finazzi","Bergamo","POLIAMBULATORIO FINAZZI SRL (cost. 2023)","04697570168","EUR 191.671 (2024); 17.580 (2023); 3 dip - reportaziende.it","Probabile SOTTO","Glauco Finazzi (AD)","035 0770751 / 340 2706511","info@poliambulatoriofinazzi.it","n/d (esiste, oscurata)","it.linkedin.com/in/glauco-finazzi","Via G.B. Berizzi 45, 24126 Bergamo","SRL neocostituita/piccola: se il business reale e' piu' grande il fatturato storico e' probabilmente in un'altra entita' - verificare"],
 ["Poliambulatorio Medivela","Torino","MEDIVELA S.R.L. (cost. 2011, ATECO 86.22.09)","10020590013","EUR 684k (2022) -> 757k (2023) -> ~826k (2024 stima); 4 dip - reportaziende.it/ufficiocamerale","DENTRO (borderline floor)","Dr. Sergio Periotto (dir. san.)","011 5612296","info@medivela.com","medivela.srl@legalmail.it","n/d","Via Vincenzo Vela 2, 10128 Torino","UNICO target Tier A con ricavi verificati vicino/dentro la soglia ICP"],
 ["Clinica Visage","Genova/Ventimiglia","Ditta indiv. / studio (ragione sociale non confermata)","02425900996 (non verificata)","n/d (non depositato)","n/d","n/d (dir. san./titolare non pubblico)","342 1915957 (anche WhatsApp)","info@clinicavisage.it","n/d","n/d","Via Ruffini 10/3, 18039 Ventimiglia; Genova (P.za De Ferrari, non confermato)","Possibile distinzione tra clinicavisage.it e 'Visage Medicina e Chir. Estetica' (P.za Piccapietra 73 Genova, 010 663351) - verificare se stessa struttura. Omonimia 'Visage Srl' San Vitaliano (NA)"],
 ["Medical Laser Clinic","Verona","MEDICAL LASER CLINIC S.R.L. (cost. 2014, ATECO 86.22.06)","04268200237","n/d (SRL con bilancio - non accessibile via web)","n/d","Dr. Stefano Anderluzzi (+ Dr. Francesco Colla)","045 990465 / 371 1457873","info@medicallaserclinic.it","n/d","n/d","Via Archimede 10, 37036 San Martino B.A. (VR)","NON confondere con 'MD Clinic Srl' (04557910231, diagnostica): il fatturato EUR 174k spesso associato via motori e' di MD Clinic, non del target"],
 ["AES Clinic","Padova","PERLA MEDICINA S.R.L. (socio unico, REA PD-447230)","05153050280","n/d (SRL con bilancio - non accessibile via web)","n/d - miglior candidata al range","Dott.ssa MianMian Wang","049 5225283 / 334 9444309","n/d (solo form sul sito)","perlamedicina@pec.it","n/d","Sede legale Via del Cristo 378, 35127 Padova; operativa P.le Stazione 7","La piu' 'aziendalizzata': SRL/poliambulatorio strutturato. Da chiudere il fatturato con visura CCIAA/Atoka"],
 ["Centro San Prospero","Bologna","Operatore da identificare ('San Prospero Srl' e' IN LIQUIDAZIONE e ad altro indirizzo)","n/d","n/d","n/d","Dr. Antonio Gotti","051 0935313 / WA 349 3684962","info@medicinaesteticasanprospero.it","n/d","n/d","Via Cesare Battisti 2/D-4/A, 40123 Bologna","Identificare la ragione sociale operativa reale (via C. Battisti 2/D) prima della visura: la 'San Prospero Srl' nota e' vecchio veicolo/omonimia"],
 ["Studio Dott.ssa Clelia Barini","Modena/Formigine","Studio prof. individuale (dal 1994)","n/d (snippet discordanti)","n/d (non depositato)","Probabile SOTTO","Dott.ssa Clelia Barini","059 557249 / 366 1391885","info@medicinaesteticabarini.it (+ dottcleliabarini@gmail.com)","n/d","n/d","Via G. Pascoli 171, 41043 Corlo di Formigine (MO) (+ La Spezia)","Studio individuale: fatturato non depositato. Docente Master Med. Estetica UniMoRe, CDA SIES"],
 ["Studio Dott.ssa Paola Molinari","Modena","Studio prof. individuale (dal 1987; OMCeO MO n.4036)","01784820365","n/d (non depositato)","n/d","Dott.ssa Paola Molinari","059 218866 / 333 2472754","medicinaesteticamolinari@gmail.com","paola.molinari.ospa@mo.omceo.it","n/d","Via Pietro Giardini 45, 41124 Modena","Studio individuale: fatturato non depositato. Teoxane Training Center, Comitato Scientifico SIES"],
]
ws6 = wb.create_sheet("Tier A - Dossier")
dh = ["Clinica","Cluster","Ragione sociale / Forma giuridica","P.IVA","Fatturato (anno / fonte)","Verdetto ICP EUR 0,8-2,5M","Referente","Telefono","Email","PEC","LinkedIn","Indirizzo","Flag / caveat"]
ws6.append(dh)
for c in range(1,len(dh)+1):
    cell = ws6.cell(1,c); cell.font=white; cell.fill=hdr_fill
    cell.alignment=Alignment(vertical="center", wrap_text=True); cell.border=border
verdict_fill = {"DENTRO (borderline floor)": PatternFill("solid", fgColor="C6E7C9"),
                "Probabile SOTTO": PatternFill("solid", fgColor="F3E1E1")}
for d in dossier:
    ws6.append(d)
for ri in range(2, ws6.max_row+1):
    for ci in range(1,len(dh)+1):
        ws6.cell(ri,ci).alignment=wrap; ws6.cell(ri,ci).border=border
    v = ws6.cell(ri,6).value
    if v in verdict_fill: ws6.cell(ri,6).fill = verdict_fill[v]
    ws6.cell(ri,1).font = Font(bold=True)
for i,w in enumerate([30,14,40,14,40,20,26,26,30,26,34,34,60],1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = "A2"
ws6.row_dimensions[1].height = 30

# reorder: put Dossier right after Sintesi
wb.move_sheet("Tier A - Dossier", -(len(wb.sheetnames)-2))

out = "/home/user/Luca/venere-target-search/output/Venere_Longlist_Target.xlsx"
wb.save(out)
print("Saved", out)
print("Totale target:", len(rows), "| Tier A:", tiers["A"], "B:", tiers["B"], "C:", tiers["C"])
print("Minori:", len(minor), "| Esclusi:", len(excluded))
